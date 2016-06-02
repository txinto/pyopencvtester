from time import gmtime, strftime, sleep
from datetime import datetime
import sqlite3
import sweepsupport as sws
import urllib
import sweepconfig

if sweepconfig.cte_export_ods:
    import pyoo
if sweepconfig.cte_export_openpyxl:
    from openpyxl import Workbook

if sweepconfig.cte_upload_web:
    import requests

# -*- coding: utf-8 -*-
"""
Automatic code generated by FoV Sweep Configurator tool
It creates a python script that commands the sweep actions and
triggers the frame acquistion shots for the configured FoV 
scanning action.
It is base (apart from python) in OpenCv.
To be successfully executed, you must provide local functions 
to give to this script the exact information needed (video source number)
and the functions to move the motors, this must included in a custom
module called "sweepsupport".
"""
if sweepconfig.cte_use_cvcam:
    import cv2

if sweepconfig.cte_use_photocam:
    import subprocess

if sweepconfig.cte_use_gphoto2:
    import gphoto2capture


def import_URL(URL):
    exec urllib.urlopen(URL).read() in globals()


# ###### Functions API ################


# ## Commands the motor to the position of that step
def commandMotor(x, y):
    if sweepconfig.cte_verbose:
        print ("Sweep step X: " + str(x) + " Y: " + str(y))
        return sws.commandMotor(x, y)


# ## Investigate if the current step has been executed
# ## you can also include here the user interaction, allowing
# ## him/her to quit the scanning operation
def stepDone():
    # Wait for command or step time
    # returns are:
    #   -1 if the sweep operation must be cancelled
    #   1 if the step has been done and the frame must be acquired
    #   0 does nothing, non blocking implementation is welcome
    return sws.stepDone()


sqlsentence = "INSERT INTO \"scan_ex_logs\" (\"step\", \"x\", \"y\", " + \
              "\"x_coord\", \"y_coord\", \"mx\", \"my\", \"mcomp\", \"mx_fdback\", \"my_fdback\", \"mcomp_fdback\", " + \
              "\"timestr\", \"scan_eng_run_id\", \"dtinit\", \"dtend\", \"created_at\", \"updated_at\") VALUES " + \
              "(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "

sqlprepare = "CREATE TABLE IF NOT EXISTS \"scan_ex_logs\" (\"id\" INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, " + \
             "\"mx\" float, \"my\" float, \"mcomp\" float, \"created_at\" datetime, \"updated_at\" datetime, " + \
             "\"scan_eng_run_id\" integer, \"step\" integer, \"x\" integer, \"y\" integer, " + \
             "\"x_coord\" float, \"y_coord\" float, \"timestr\" varchar, \"dtinit\" datetime, \"dtend\" datetime, " + \
             "\"mx_fdback\" float, \"my_fdback\" float, \"mcomp_fdback\" float);"

sqlsentence2 = "INSERT INTO \"scan_eng_runs\" (\"name\", \"max_l1_speed\", \"max_l2_speed\", \"max_l3_speed\", " + \
               "\"created_at\", \"updated_at\", \"scan_ex_id\") VALUES (?, ?, ?, ?, ?, ?, ?) "

sqlprepare2 = "CREATE TABLE IF NOT EXISTS \"scan_eng_runs\" (\"id\" INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, " + \
              "\"name\" varchar, \"max_l1_speed\" float, \"max_l2_speed\" float, \"max_l3_speed\" float, " + \
              "\"created_at\" datetime, \"updated_at\" datetime, \"scan_ex_id\" integer);"

firstDbSentence = True
sweep_eng_run_id = None

engrunsheet = None
exlogsheet = None
doc = None
docrow = 0


def dbinsert(dbcon, cur_step, step_x, step_y, step_x_coord, step_y_coord, mx_setpoint, my_setpoint, mcomp_setpoint,
             mx_pos, my_pos, mcomp_pos, timestamp, dt_init, dt_end, ex_id, run_id):
    global firstDbSentence
    global docrow

    sdtcam = str(dt_end)
    sdtinit = str(dt_init)

    if firstDbSentence:
        item = [timestamp, 100, 100, 100, dt_end, dt_end, ex_id]
        print(sqlsentence2)
        dbcon.execute(sqlsentence2, item)
        run_id = dbcon.lastrowid
        firstDbSentence = False
        if sweepconfig.cte_export_ods:
            engrunsheet[0, 0:8].values = ["id", "name", "max_l1_speed", "max_l2_speed", "max_l3_speed", "created_at",
                                          "updated_at", "scan_ex_id"]
            engrunsheet[1, 0].value = run_id
            item = [timestamp, 100, 100, 100, sdtcam, sdtcam, ex_id]
            engrunsheet[1, 1:8].values = item
            exlogsheet[0, 0:18].values = ["id", "step", "x", "y", "x_coord", "y_coord", "mx", "my", "mcomp",
                                          "mx_fdback", "my_fdback", "mcomp_fdback", "timestr", "scan_eng_run_id",
                                          "dtinit", "dtend", "created_at", "updated_at"]
        if sweepconfig.cte_export_openpyxl:
            idx = 0
            item = ["id", "name", "max_l1_speed", "max_l2_speed", "max_l3_speed", "created_at",
                    "updated_at", "scan_ex_id"]
            for row in engrunsheet.iter_rows('A1:H1'):
                for cell in row:
                    cell.value = item[idx]
                    idx += 1
            engrunsheet['A2'] = run_id
            idx = 0
            item = [timestamp, 100, 100, 100, sdtcam, sdtcam, ex_id]
            for row in engrunsheet.iter_rows('B2:H2'):
                for cell in row:
                    print("idx: " + str(idx))
                    cell.value = item[idx]
                    idx += 1

            item = ["id", "step", "x", "y", "x_coord", "y_coord", "mx", "my", "mcomp",
                    "mx_fdback", "my_fdback", "mcomp_fdback", "timestr", "scan_eng_run_id",
                    "dtinit", "dtend", "created_at", "updated_at"]
            idx = 0
            for row in exlogsheet.iter_rows('A1:R1'):
                for cell in row:
                    cell.value = item[idx]
                    idx += 1

        # Web information upload
        if sweepconfig.cte_upload_web:
            postdata = {'sweep_eng_run': {"name": timestamp, "max_l1_speed": "33"}}
            r = requests.post(sweepconfig.cte_web_root + "/scan_eng_runs", data=postdata)
            print(r.status_code, r.reason)

        docrow += 1

    item = [cur_step, step_x, step_y, step_x_coord, step_y_coord, mx_setpoint, my_setpoint, mcomp_setpoint, mx_pos,
            my_pos, mcomp_pos, timestamp, run_id, dt_init, dt_end, dt_end, dt_end]
    dbcon.execute(sqlsentence, item)

    if sweepconfig.cte_export_ods:
        item = [cur_step, step_x, step_y, step_x_coord, step_y_coord, mx_setpoint, my_setpoint, mcomp_setpoint, mx_pos,
                my_pos, mcomp_pos, timestamp, run_id, sdtinit, sdtcam, sdtcam, sdtcam]
        exlogsheet[docrow, 0].value = dbcon.lastrowid
        exlogsheet[docrow, 1:18].values = item

    if sweepconfig.cte_export_openpyxl:
        item = [cur_step, step_x, step_y, step_x_coord, step_y_coord, mx_setpoint, my_setpoint, mcomp_setpoint, mx_pos,
                my_pos, mcomp_pos, timestamp, run_id, sdtinit, sdtcam, sdtcam, sdtcam]
        exlogsheet['A' + str(docrow + 1)] = dbcon.lastrowid
        idx = 0
        for row in exlogsheet.iter_rows('B' + str(docrow + 1) + ':R' + str(docrow + 1)):
            for cell in row:
                cell.value = item[idx]
                idx += 1

    docrow += 1

    return run_id


def dbprepare(dbcon):
    global engrunsheet
    global exlogsheet
    global doc

    dbcon.execute(sqlprepare2)
    dbcon.execute(sqlprepare)
    if sweepconfig.cte_export_ods:
        oodesktop = pyoo.Desktop('localhost', 2002)
        doc = oodesktop.create_spreadsheet()
        engrunsheet = doc.sheets.create('EngRun', index=1)
        exlogsheet = doc.sheets.create('ExLog', index=1)
        del doc.sheets[0]
    if sweepconfig.cte_export_openpyxl:
        doc = Workbook()
        engrunsheet = doc.create_sheet()
        engrunsheet.title = 'EngRun'
        exlogsheet = doc.create_sheet()
        exlogsheet.title = 'EngLog'
        ws = doc.active
        doc.remove_sheet(ws)
    return True


# ##### END Functions ############

# ##### Automatically generated code ###########

sweep_ex_id = 2
steps = [{'c': (0), 'x': (0), 'y': (0), 'x_coord': (0.0), 'y_coord': (0.0)},
         {'c': (1), 'x': (1), 'y': (0), 'x_coord': (0.0018), 'y_coord': (0.0)},
         {'c': (2), 'x': (1), 'y': (1), 'x_coord': (0.0018), 'y_coord': (0.0008)},
         {'c': (3), 'x': (0), 'y': (1), 'x_coord': (0.0), 'y_coord': (0.0008)},
         {'c': (4), 'x': (-1), 'y': (1), 'x_coord': (-0.0018), 'y_coord': (0.0008)},
         {'c': (5), 'x': (-1), 'y': (0), 'x_coord': (-0.0018), 'y_coord': (0.0)},
         {'c': (6), 'x': (-1), 'y': (-1), 'x_coord': (-0.0018), 'y_coord': (-0.0008)},
         {'c': (7), 'x': (0), 'y': (-1), 'x_coord': (0.0), 'y_coord': (-0.0008)},
         {'c': (8), 'x': (1), 'y': (-1), 'x_coord': (0.0018), 'y_coord': (-0.0008)},
         {'c': (9), 'x': (2), 'y': (-1), 'x_coord': (0.0036), 'y_coord': (-0.0008)},
         {'c': (10), 'x': (2), 'y': (0), 'x_coord': (0.0036), 'y_coord': (0.0)},
         {'c': (11), 'x': (2), 'y': (1), 'x_coord': (0.0036), 'y_coord': (0.0008)},
         {'c': (12), 'x': (2), 'y': (2), 'x_coord': (0.0036), 'y_coord': (0.0016)},
         {'c': (13), 'x': (1), 'y': (2), 'x_coord': (0.0018), 'y_coord': (0.0016)},
         {'c': (14), 'x': (0), 'y': (2), 'x_coord': (0.0), 'y_coord': (0.0016)},
         {'c': (15), 'x': (-1), 'y': (2), 'x_coord': (-0.0018), 'y_coord': (0.0016)},
         {'c': (16), 'x': (-2), 'y': (2), 'x_coord': (-0.0036), 'y_coord': (0.0016)},
         {'c': (17), 'x': (-2), 'y': (1), 'x_coord': (-0.0036), 'y_coord': (0.0008)},
         {'c': (18), 'x': (-2), 'y': (0), 'x_coord': (-0.0036), 'y_coord': (0.0)},
         {'c': (19), 'x': (-2), 'y': (-1), 'x_coord': (-0.0036), 'y_coord': (-0.0008)},
         {'c': (20), 'x': (-2), 'y': (-2), 'x_coord': (-0.0036), 'y_coord': (-0.0016)},
         {'c': (21), 'x': (-1), 'y': (-2), 'x_coord': (-0.0018), 'y_coord': (-0.0016)},
         {'c': (22), 'x': (0), 'y': (-2), 'x_coord': (0.0), 'y_coord': (-0.0016)},
         {'c': (23), 'x': (1), 'y': (-2), 'x_coord': (0.0018), 'y_coord': (-0.0016)},
         {'c': (24), 'x': (2), 'y': (-2), 'x_coord': (0.0036), 'y_coord': (-0.0016)}, ]

# ##### Automatically generated steps table ###########
# ##### END Automatically generated code ###########

# ### START EXECUTION ######

# Prepare the scan loop
curStep = 0
done = 0
# Create timestamp
timestr = strftime("%Y%m%d%H%M%S", gmtime())

if sweepconfig.cte_use_cvcam:
    # Cam has the video source
    cam = cv2.VideoCapture(sweepconfig.cte_camsource)
    if sweepconfig.cte_verbose:
        print ("Camera resolution:")
        print ("* Horizontal: " + str(cam.get(cv2.CAP_PROP_FRAME_WIDTH)))
        print ("* Vertical: " + str(cam.get(cv2.CAP_PROP_FRAME_HEIGHT)))

        # subprocess.check_call("exit 1", shell=True)

if sweepconfig.cte_disable_motors_first:
    sws.disableMotors()

if sweepconfig.cte_enable_motors_first:
    sws.enableMotors()

if sweepconfig.cte_reset_motors_first:
    sws.resetMotors()
    print("Check motor positions after resets")
    sleep(sweepconfig.cte_stabilization_time)
    sws.motorPositions()
else:
    print("Check initial motor positions")
    sleep(sweepconfig.cte_stabilization_time)
    sws.motorPositions()

# Prepare the Database
db = sqlite3.connect('./db/log.sqlite3')
if not db:
    ret = False
else:
    ret = True
    dbc = db.cursor()
    ret = dbprepare(dbc)

if not ret:
    done = -1
    print "Database ERROR! Aborting"

# Steps loop
# until ESC key is pressed
# or steps have finished
endStep = len(steps)
# endStep = 4
while (done != -1) and (curStep < endStep):
    # In stepX and stepY we have the step positions to be done
    stepX = steps[curStep]['x']
    stepY = steps[curStep]['y']
    stepXcoord = steps[curStep]['x_coord']
    stepYcoord = steps[curStep]['y_coord']
    # Command motor position for this step
    dtinit = datetime.now()
    done, mx, my, mcomp = commandMotor(stepXcoord, stepYcoord)
    # Wait command to end
    while done == 0:
        done = stepDone()
    # END Command motor position for this step    
    if done != -1:
        # Acquire image
        dtcam = datetime.now()
        capture_done = False
        if sweepconfig.cte_use_cvcam:
            ret, frame = cam.read()
            # save to disk
            strg = sweepconfig.cte_fileprefix + '%s_%03d_%03d.png' % (timestr, sweep_ex_id, curStep)
            cv2.imwrite(sweepconfig.cte_framePath + strg, frame)
            # show the image
            cv2.imshow('Current Frame', frame)
            capture_done = True
        if sweepconfig.cte_use_photocam:
            # We configure the image capture
            strg = 'D%s_%03d_%03d.jpg' % (timestr, sweep_ex_id, curStep)
            cmd = sweepconfig.cte_cameractrl_path + sweepconfig.cte_cameractrl_command
            args = sweepconfig.cte_cameractrl_filenamecmd + " " + strg
            print("Photo set filename: " + cmd + " " + args)
            subprocess.check_output([cmd, args])
            args = sweepconfig.cte_cameractrl_capturecmd
            print("Photo capture frame: " + cmd + " " + args)
            subprocess.check_output([cmd, args])
            capture_done = True
        if sweepconfig.cte_use_gphoto2:
            strg = sweepconfig.cte_gphoto2_filename_root + '%s_%03d_%03d.jpg' % (timestr, sweep_ex_id, curStep)
            gphoto2capture.capture(sweepconfig.cte_gphoto2_framePath, strg, False)
            capture_done = True
        if not capture_done:
            # Wait some ms to stabilyze before reading position
            # not necessary if capture has been taken
            sleep(sweepconfig.cte_stabilization_time)

        # acquire the motor status
        mx_fdback, my_fdback, mcomp_fdback = sws.motorPositions()
        print ("Motor | mx: " + str(mx_fdback) + ", my: " + str(my_fdback) + ", mcomp: " + str(mcomp_fdback))
        # BD information store
        sweep_eng_run_id = dbinsert(dbc, curStep, stepX, stepY, stepXcoord, stepYcoord, mx, my, mcomp, mx_fdback,
                                    my_fdback, mcomp_fdback, timestr, dtinit, dtcam, sweep_ex_id, sweep_eng_run_id)
        curStep += 1

# End of program, steps have finished or someone has cancelled the scan process
if curStep < len(steps) and sweepconfig.cte_verbose:
    # Scan process was cancelled
    print ("Scan process was cancelled")
    dummy = 0  # Dummy for avoiding indentation failures

db.commit()
db.close()

if sweepconfig.cte_export_ods:
    doc.save("./db/" + timestr + ".ods")
    doc.close()

if sweepconfig.cte_export_openpyxl:
    doc.save("./db/" + timestr + ".ods")

if sweepconfig.cte_use_cvcam:
    cam.release()
    cv2.destroyAllWindows()

sws.motorClose()